import openpyxl

def read_txt():
	with open(r'DKBanword.txt','r',encoding='utf-8') as f:#打开txt文件
		data = f.read().replace('\n', '')
		#遍历txt文件内容存放到列表
		contents = data.split('、')
		#过滤重复字段、空字符字段
		newContents=[]
		for c in contents:
			if c not in newContents and len(c)!=0:
				newContents.append(c)
	return newContents

def write_excel(contents):
	wb=openpyxl.Workbook()#创建1个工作簿
	ws=wb.create_sheet(u'Data_L')#用工作簿去创建工作表sheet
	ws.cell(1,1,u'/备注')
	ws.cell(1,2,u'禁语')
	ws.cell(2,2,u'string')
	ws.cell(3,2,u'banWord')
	for i,content in enumerate(contents):
		cell = ws.cell(row=i+4, column=2)
		cell.value = content#用工作表sheet调用单元格，写入内容
	wb.save(r"DKBanword.xlsx")#保存文件名		

contents = read_txt()
write_excel(contents)